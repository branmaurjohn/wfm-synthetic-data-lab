from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl

# Local imports (works once installed via -e . OR when repo root is on sys.path)
from wfm_synth.config import load_config
from wfm_synth.schema_registry import snapshot_path_for_table, load_snapshot
from wfm_synth.profiles import load_profile
from wfm_synth.column_mapper import build_mapping

from wfm_synth.generate_timecard_total import generate_vtimecardtotal
from wfm_synth.generate_accrual_balance import generate_vaccrualbalance
from wfm_synth.generate_dim_business_structure import generate_vdimbusinessstructure


def _cmd_doctor(_: argparse.Namespace) -> int:
    repo_root = Path(__file__).resolve().parents[1]
    print(f"Repo root: {repo_root}")

    needed = [
        repo_root / "schemas" / "snapshots" / "vTimecardTotal.schema.json",
        repo_root / "schemas" / "snapshots" / "vAccrualBalance.schema.json",
        repo_root / "schemas" / "snapshots" / "vDimBusinessStructure.schema.json",
        repo_root / "scenarios" / "baptist_south_fl.yaml",
    ]
    missing = [p for p in needed if not p.exists()]
    if missing:
        print("MISSING FILES:")
        for p in missing:
            print(f"  - {p}")
        return 2

    try:
        import faker  # noqa: F401
        import pandas  # noqa: F401
        import yaml  # noqa: F401
        import openpyxl  # noqa: F401
    except Exception as e:
        print("Python dependency problem:", repr(e))
        return 3

    print("OK: repo layout + key dependencies are good")
    return 0


def _is_table_header(name, unique_identifier, dtype) -> bool:
    if not name:
        return False
    name = str(name)
    return name.startswith("v") and unique_identifier and (dtype is None or str(dtype).strip() == "")


def _cmd_extract_schema(args: argparse.Namespace) -> int:
    xlsx_path = Path(args.xlsx).expanduser().resolve()
    table_name = args.table
    out_path = Path(args.out).expanduser().resolve()

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append([ws.cell(r, c).value for c in range(1, 8)])

    start = None
    unique_identifier = None
    for i, row in enumerate(rows):
        name, uniq, dtype = row[0], row[1], row[2]
        if name == table_name and _is_table_header(name, uniq, dtype):
            start = i
            unique_identifier = str(uniq) if uniq else None
            break

    if start is None:
        print(f"Table not found in dictionary: {table_name}")
        return 2

    cols = []
    for j in range(start + 1, len(rows)):
        name, uniq, dtype, desc, source, pii, scrub = rows[j]
        if _is_table_header(name, uniq, dtype):
            break
        if not name or not dtype:
            continue
        cols.append(
            {
                "name": str(name),
                "dtype": str(dtype),
                "description": str(desc) if desc else None,
                "pii": str(pii) if pii else None,
                "scrubbing": str(scrub) if scrub else None,
            }
        )

    payload = {"table": table_name, "unique_identifier": unique_identifier, "columns": cols}
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    print(f"Wrote schema snapshot: {out_path}")
    print(f"Columns: {len(cols)}")
    return 0


def _cmd_generate(args: argparse.Namespace) -> int:
    scenario = Path(args.scenario).resolve()
    table = args.table
    out_dir = Path(args.out_dir).resolve()

    cfg = load_config(str(scenario))
    snap_path = snapshot_path_for_table(table)
    snap = load_snapshot(snap_path)

    if snap.table != table:
        print(f"Snapshot table mismatch: expected {table}, got {snap.table}")
        return 2

    schema_cols = snap.column_names

    profile_cols = None
    mapping = None
    try:
        profile = load_profile(table)
        profile_cols = profile.columns
        mapping = build_mapping(profile_cols)
    except Exception:
        profile_cols = None
        mapping = None

    if table == "vTimecardTotal":
        out = generate_vtimecardtotal(
            cfg,
            {"table": snap.table, "unique_identifier": snap.unique_identifier, "columns": snap.columns},
            out_dir,
            profile_cols,
            mapping,
        )
    elif table == "vAccrualBalance":
        out = generate_vaccrualbalance(cfg, schema_cols, out_dir, profile_cols, mapping)
    elif table == "vDimBusinessStructure":
        out = generate_vdimbusinessstructure(cfg, schema_cols, out_dir, profile_cols, mapping)
    else:
        print(f"No generator registered for table: {table}")
        return 2

    print(f"Wrote: {out}")
    print(f"Metadata: {out_dir / 'run_metadata.json'}")
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(prog="wfm-synth", description="WFM Synthetic Data Lab CLI")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_doc = sub.add_parser("doctor", help="Validate repo layout and Python deps")
    p_doc.set_defaults(func=_cmd_doctor)

    p_es = sub.add_parser("extract-schema", help="Extract a table schema snapshot from Omni dictionary xlsx")
    p_es.add_argument("xlsx", help="Path to Omni Data Hub Data Dictionary xlsx")
    p_es.add_argument("table", help="Table name like vTimecardTotal")
    p_es.add_argument("out", help="Output json path for schema snapshot")
    p_es.set_defaults(func=_cmd_extract_schema)

    p_gen = sub.add_parser("generate", help="Generate a table dataset for a scenario")
    p_gen.add_argument("scenario", help="Scenario YAML path")
    p_gen.add_argument("table", help="Table name like vTimecardTotal")
    p_gen.add_argument("out_dir", help="Output folder")
    p_gen.set_defaults(func=_cmd_generate)

    args = parser.parse_args()
    rc = args.func(args)
    raise SystemExit(rc)
