import json
import sys
from pathlib import Path
import openpyxl

def is_table_header(name, unique_identifier, dtype):
    if not name:
        return False
    name = str(name)
    # Table header rows look like: vSomething + unique identifier filled + dtype blank
    return name.startswith("v") and unique_identifier and (dtype is None or str(dtype).strip() == "")

def extract_table_schema(xlsx_path: Path, table_name: str, out_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # rows: columns A-G
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append([ws.cell(r, c).value for c in range(1, 8)])

    # find table header
    start = None
    unique_identifier = None
    for i, row in enumerate(rows):
        name, uniq, dtype = row[0], row[1], row[2]
        if name == table_name and is_table_header(name, uniq, dtype):
            start = i
            unique_identifier = str(uniq) if uniq else None
            break

    if start is None:
        raise SystemExit(f"Table not found in dictionary: {table_name}")

    cols = []
    for j in range(start + 1, len(rows)):
        name, uniq, dtype, desc, source, pii, scrub = rows[j]
        if is_table_header(name, uniq, dtype):
            break
        if not name or not dtype:
            continue

        cols.append({
            "name": str(name),
            "dtype": str(dtype),
            "description": str(desc) if desc else None,
            "pii": str(pii) if pii else None,
            "scrubbing": str(scrub) if scrub else None,
        })

    payload = {
        "table": table_name,
        "unique_identifier": unique_identifier,
        "columns": cols
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    print(f"Wrote schema snapshot: {out_path}")
    print(f"Columns: {len(cols)}")

def main():
    if len(sys.argv) != 4:
        raise SystemExit("Usage: python tools/extract_schema.py <xlsx_path> <table_name> <out_json_path>")

    xlsx_path = Path(sys.argv[1]).expanduser().resolve()
    table_name = sys.argv[2]
    out_path = Path(sys.argv[3]).expanduser().resolve()

    extract_table_schema(xlsx_path, table_name, out_path)

if __name__ == "__main__":
    main()

